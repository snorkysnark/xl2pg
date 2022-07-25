from argparse import ArgumentParser
import json
from pathlib import Path
from typing import TypedDict
import textwrap
import sys

import openpyxl
import psycopg
from psycopg.sql import SQL, Identifier, Placeholder


class MappingConfig(TypedDict):
    target_schema: str
    target_table: str
    sheet_number: int
    skip_rows: int
    mappings: dict[str, int]


def printerr(*args, **kwargs):
    print(*args, **kwargs, file=sys.stderr)


def upload(
    conn: psycopg.Connection,
    wb: openpyxl.Workbook,
    map_cfg: MappingConfig,
    clear: bool = False,
):
    table_path = SQL("{}.{}").format(
        Identifier(map_cfg["target_schema"], map_cfg["target_table"])
    )
    insert_statement = SQL(
        textwrap.dedent(
            """\
            INSERT INTO {table}(
                {fields}
            ) VALUES ({values})"""
        )
    ).format(
        table=table_path,
        fields=SQL(",\n").join(map(Identifier, map_cfg["mappings"].keys())),
        values=SQL(", ").join([Placeholder()] * len(map_cfg["mappings"])),
    )
    printerr("Generated statement:")
    printerr(insert_statement.as_string(conn))
    printerr()

    sheet = wb.worksheets[map_cfg["sheet_number"]]

    with conn.cursor() as cursor:
        if clear:
            printerr(f"Clearing {table_path.as_string(cursor)}")
            cursor.execute(SQL("DELETE FROM {}").format(table_path))

        max_row = sheet.max_row
        for row_index in range(map_cfg["skip_rows"] + 1, max_row + 1):
            output_row = []

            for column_index in map_cfg.values():
                value = sheet.cell(row_index, column_index).value
                output_row.append(value)
            printerr(
                "Inserting row", row_index, "out of", max_row, end="\r", flush=True
            )

            cursor.execute(insert_statement, output_row, prepare=True)
        printerr()


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("spreadsheet", type=Path, help="xlsx file to upload")
    parser.add_argument(
        "-d",
        "--db",
        type=Path,
        required=True,
        help='Database config as a json file: { "dbname": "", "user": "", "password": "" }',
    )
    parser.add_argument(
        "-m",
        "--map",
        required=True,
        type=Path,
        help="Json mapping from excel fields to postgres table",
    )
    parser.add_argument(
        "--clear", action="store_true", help="perform DELETE FROM target table"
    )
    args = parser.parse_args()

    conninfo = json.loads(args.db.read_text())
    map_cfg: MappingConfig = json.loads(args.map.read_text())

    printerr("Loading spreadsheet")
    wb = openpyxl.load_workbook(args.spreadsheet)

    printerr("Connecting to database")
    with psycopg.connect(**conninfo) as conn:
        upload(conn=conn, wb=wb, map_cfg=map_cfg, clear=args.clear)
